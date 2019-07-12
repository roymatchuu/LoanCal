import openpyxl
import os

from Loan import *
from datetime import datetime

# method for calculating the number of days for subsidized a subsidized loan
def calculateDays(loan):
    today = datetime.today().strftime("%m/%d/%y")
    borDate = datetime.strftime(loan.dateBorrowed.value, "%m/%d/%y")

    a = datetime.strptime(borDate, "%m/%d/%y")
    b = datetime.strptime(today, "%m/%d/%y")

    dTime = b - a
    return b - a


# method for calculating the loan
def loanCal(l):
    amount = 0

    print("Please answer the next question using Y or N.\n")
    inSchool = input("Are you currently in school? ")

    for i in range(len(l)):
        loanType = l[i].type.value

        if inSchool == 'Y' or inSchool == 'y':

            if loanType[0] == 'S':
                # print("In Subsidized if")
                # print("added value: ", l[i].amount.value)
                amount += l[i].amount.value
            elif loanType[0] == 'U':

                temp = calculateDays(l[i])
                numDays = temp.days

                tempAmt = (l[i].interest.value/36500) * l[i].amount.value
                # print("accrued interest: ", l[i].interest.value/36500)
                # print("tempAmt: ", tempAmt)

                # print(numDays)
                # print("accrued interest: ", tempAmt * numDays)
                # print("testVal: ", l[i].amount.value + (tempAmt * numDays))
                # # print("t2: ", tempAmt * numDays)
                # # print(amount + (tempAmt * numDays))


                # print(l[i].name.value, ": ", l[i].amount.value)

                # print("added unsub value: ", (tempAmt * numDays) + l[i].amount.value)
                amount += (tempAmt * numDays) + l[i].amount.value

    print("Your total amount of loans is approximately worth $", "%.2f" % amount)


def main():
    os.chdir('D:\Documents - HDD')
    wb = openpyxl.load_workbook('Loan_Spreadsheet.xlsx')
    type(wb)

    # changing the sheet to the "List" sheets where the loans are
    wb.active = 1
    sheet = wb.active
    loanList = []

    # creates the "Loan" object and appends to a list
    for x in range(4):
        name = sheet.cell(x + 3, 1)
        t = sheet.cell(x + 3, 2)
        amt = sheet.cell(x + 3, 3)
        interest = sheet.cell(x + 3, 4)
        dateBorrowed = sheet.cell(x + 3, 5)

        print(name.value, " | ", t.value, " | ", amt.value, " | ", interest.value, " | ", dateBorrowed.value)
        l = Loan(name, t, amt, interest, dateBorrowed)
        loanList.append(l)

    print("\n")
    loanCal(loanList)


if __name__ == "__main__":
    main()
